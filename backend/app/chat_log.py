"""Durable log of every question a visitor asks, exportable as an Excel sheet.

Two storage layers, and which one is in charge depends on configuration:

* **Turso (libSQL)**, when TURSO_DATABASE_URL and TURSO_AUTH_TOKEN are set. This
  is the one that matters in production. The host wipes the container filesystem
  on every deploy and every cold start, so a file-backed log quietly loses
  everything a few times a day; a database does not.
* **A JSONL file**, otherwise. One question per line, appended the moment the
  answer finishes. Used for local development, and as the write-ahead buffer
  underneath Turso.

Writes always hit the file first and are mirrored to Turso by a background
writer, because `append` is called from the `finally` of a cancelled SSE
generator: it has to be non-blocking and it has to not raise. Reads come from
Turso whenever it is configured, since the file is the copy that disappears.

Only the owner (ADMIN_TOKEN) can read or clear this log; visitors can neither
see it nor delete from it.
"""

import csv
import io
import json
import logging
import queue
import re
import threading
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterator, Optional

from .turso import TursoClient, TursoError

logger = logging.getLogger("ai-portfolio.chatlog")

IST = timezone(timedelta(hours=5, minutes=30))

try:  # optional — without it the export falls back to CSV, which Excel opens fine
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    XLSX_AVAILABLE = True
except ImportError:  # pragma: no cover - depends on the deployment image
    XLSX_AVAILABLE = False


# (header, entry key, column width) — the order here is the order in the sheet.
# Who asked comes before what they asked, because that is what the sheet is for.
COLUMNS: list[tuple[str, str, int]] = [
    ("#", "row", 6),
    ("Date (IST)", "date_ist", 12),
    ("Time (IST)", "time_ist", 10),
    ("Type", "kind", 11),
    ("Name", "name", 18),
    ("Email", "email", 26),
    ("Company / role", "company", 22),
    ("LinkedIn", "linkedin", 34),
    ("Question", "question", 60),
    ("Answer", "answer", 70),
    ("IP address", "ip", 16),
    ("City", "city", 14),
    ("Region", "region", 14),
    ("Country", "country", 12),
    ("Network / ISP", "isp", 22),
    ("Device", "device", 10),
    ("OS", "os", 14),
    ("Browser", "browser", 14),
    ("Page", "page", 30),
    ("Came from", "referrer", 30),
    ("Their timezone", "timezone", 20),
    ("Screen", "screen", 12),
    ("Browser language", "browser_language", 16),
    ("Reply language", "language", 12),
    ("Visitor ID", "visitor_id", 20),
    ("Session ID", "session_id", 20),
    ("Conversation", "conversation_id", 20),
    ("Turn", "turn", 6),
    ("Model used", "model", 26),
    ("Status", "status", 10),
    ("Flagged term", "flagged_term", 14),
    ("Reply time (s)", "duration_s", 12),
    ("User agent", "user_agent", 40),
    ("Asked at (UTC)", "asked_at", 22),
]

KIND_LABELS = {"chat": "Chat", "abuse": "Flagged"}


SCHEMA = [
    """CREATE TABLE IF NOT EXISTS chat_log (
        id       TEXT PRIMARY KEY,
        asked_at TEXT NOT NULL,
        payload  TEXT NOT NULL
    )""",
    "CREATE INDEX IF NOT EXISTS chat_log_asked_at ON chat_log (asked_at)",
    """CREATE TABLE IF NOT EXISTS chat_meta (
        key   TEXT PRIMARY KEY,
        value TEXT NOT NULL
    )""",
    """CREATE TABLE IF NOT EXISTS strikes (
        key      TEXT PRIMARY KEY,
        strikes  INTEGER NOT NULL,
        last_at  REAL NOT NULL,
        term     TEXT,
        question TEXT
    )""",
]


class ChatLogStore:
    """Append-only question log, with owner-only read/export/clear."""

    def __init__(
        self, path: Path, max_rows: int = 20000, turso: Optional[TursoClient] = None
    ) -> None:
        self._path = path
        # Sidecar, so it survives the log file being deleted — that is the whole
        # point of it. Holds when the log was last wiped.
        self._meta_path = path.with_name(path.stem + ".meta.json")
        self._max_rows = max_rows
        self._lock = threading.Lock()

        self._turso = turso if (turso and turso.enabled) else None
        self._queue: "queue.Queue[Optional[dict[str, Any]]]" = queue.Queue()
        self._writer: Optional[threading.Thread] = None
        if self._turso and not self._init_turso():
            self._turso = None  # unreachable at boot — carry on with the file

        self._count = self._remote_count() if self._turso else self._count_lines()
        self._reset_at = self._read_reset_at()
        logger.info(
            "Chat log on %s (%d entries)",
            "Turso" if self._turso else f"file {self._path}",
            self._count,
        )

    # -- Turso ------------------------------------------------------------

    def _init_turso(self) -> bool:
        """Create the tables and start the writer. False if the database is
        unreachable, which downgrades the log to file-only rather than 500ing
        every chat."""
        try:
            self._turso.batch((sql, ()) for sql in SCHEMA)
        except TursoError:
            logger.exception("Turso is configured but unreachable — falling back to the file")
            return False

        self._writer = threading.Thread(
            target=self._drain, name="chat-log-writer", daemon=True
        )
        self._writer.start()
        return True

    def _drain(self) -> None:
        """Mirror queued entries into Turso, one round trip per batch.

        Runs off the request path entirely: a slow database delays the row
        landing in the table, never the visitor's answer.
        """
        while True:
            entry = self._queue.get()
            if entry is None:  # shutdown sentinel
                self._queue.task_done()
                return

            batch = [entry]
            while len(batch) < 50:  # opportunistic: whatever else is waiting
                try:
                    extra = self._queue.get_nowait()
                except queue.Empty:
                    break
                if extra is None:
                    self._queue.put(None)  # put the sentinel back for the next loop
                    break
                batch.append(extra)

            try:
                self._turso.batch(
                    (
                        "INSERT OR REPLACE INTO chat_log (id, asked_at, payload) VALUES (?, ?, ?)",
                        (
                            row.get("id") or f"{time.time()}",
                            row.get("asked_at") or "",
                            json.dumps(row, ensure_ascii=False, default=str),
                        ),
                    )
                    for row in batch
                )
            except TursoError:
                logger.exception("Could not mirror %d row(s) to Turso", len(batch))

            for _ in batch:
                self._queue.task_done()

    def _remote_count(self) -> int:
        try:
            rows = self._turso.execute("SELECT COUNT(*) AS n FROM chat_log")
            return int(rows[0]["n"]) if rows else 0
        except (TursoError, KeyError, ValueError):
            logger.exception("Could not count the Turso chat log")
            return 0

    def flush(self, timeout: float = 5.0) -> None:
        """Wait for queued rows to reach Turso. Called on shutdown."""
        if not self._turso:
            return
        finished = threading.Thread(target=self._queue.join, daemon=True)
        finished.start()
        finished.join(timeout)

    # -- writing -----------------------------------------------------------

    def append(self, entry: dict[str, Any]) -> None:
        """Persist one question. Never raises — a logging failure must not
        break the answer the visitor is already reading."""
        try:
            line = json.dumps(entry, ensure_ascii=False, default=str)
            with self._lock:
                self._path.parent.mkdir(parents=True, exist_ok=True)
                with self._path.open("a", encoding="utf-8") as handle:
                    handle.write(line + "\n")
                self._count += 1
                over = self._max_rows and self._count > self._max_rows and not self._turso
            if self._turso:
                self._queue.put(entry)
            if over:
                self._trim()
        except Exception:  # noqa: BLE001 — logging is best-effort by design
            logger.exception("Could not append to the chat log")

    def _trim(self) -> None:
        """Drop the oldest entries once the file passes its ceiling.

        Set CHAT_LOG_MAX_ROWS=0 to keep everything forever instead.
        """
        try:
            with self._lock:
                lines = self._path.read_text(encoding="utf-8").splitlines()
                keep = lines[-self._max_rows :]
                self._path.write_text("\n".join(keep) + "\n", encoding="utf-8")
                dropped = len(lines) - len(keep)
                self._count = len(keep)
            logger.warning("Chat log hit %d rows — dropped the %d oldest", self._max_rows, dropped)
        except Exception:  # noqa: BLE001
            logger.exception("Could not trim the chat log")

    # -- reading -----------------------------------------------------------

    # -- the reset marker --------------------------------------------------
    #
    # Wiping the log has to reach the visitors too: their name and email live in
    # their own browser, and a server-side delete cannot touch that. So the wipe
    # stamps a timestamp here, every client compares it against the one it last
    # saw, and anything newer means "forget who you are and ask again".
    #
    # A timestamp rather than a counter, and "newer wins" rather than "different
    # wins", so that losing this file — which Render's free plan does on every
    # cold start — reads as 0 and quietly resets nobody.

    def _read_reset_at(self) -> float:
        if self._turso:
            try:
                rows = self._turso.execute(
                    "SELECT value FROM chat_meta WHERE key = ?", ("reset_at",)
                )
                return float(rows[0]["value"]) if rows else 0.0
            except (TursoError, KeyError, ValueError, TypeError):
                logger.exception("Could not read the reset marker from Turso")
                return 0.0
        try:
            return float(json.loads(self._meta_path.read_text(encoding="utf-8"))["reset_at"])
        except (OSError, ValueError, KeyError, TypeError):
            return 0.0

    def _write_reset_at(self, value: float) -> None:
        if self._turso:
            try:
                self._turso.execute(
                    "INSERT OR REPLACE INTO chat_meta (key, value) VALUES (?, ?)",
                    ("reset_at", str(value)),
                )
                return
            except TursoError:
                logger.exception("Could not persist the reset marker to Turso")
                return
        try:
            self._meta_path.parent.mkdir(parents=True, exist_ok=True)
            self._meta_path.write_text(
                json.dumps({"reset_at": value}), encoding="utf-8"
            )
        except OSError:
            logger.exception("Could not persist the chat-log reset marker")

    @property
    def reset_at(self) -> float:
        """Unix seconds of the last wipe; 0 if it has never been wiped."""
        return self._reset_at

    def _count_lines(self) -> int:
        if not self._path.exists():
            return 0
        try:
            with self._path.open("r", encoding="utf-8") as handle:
                return sum(1 for line in handle if line.strip())
        except OSError:
            return 0

    def _iter_entries(self) -> Iterator[dict[str, Any]]:
        if not self._path.exists():
            return
        try:
            with self._path.open("r", encoding="utf-8") as handle:
                for line in handle:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        yield json.loads(line)
                    except json.JSONDecodeError:
                        continue  # a half-written line from a killed process
        except OSError:
            logger.exception("Could not read the chat log")

    @property
    def count(self) -> int:
        return self._count

    @property
    def path(self) -> Path:
        return self._path

    def all_entries(self, newest_first: bool = False) -> list[dict[str, Any]]:
        if self._turso:
            entries = self._remote_entries()
        else:
            entries = list(self._iter_entries())
        return entries[::-1] if newest_first else entries

    def _remote_entries(self) -> list[dict[str, Any]]:
        """Oldest first, to match the file order every caller already expects."""
        try:
            rows = self._turso.execute(
                "SELECT payload FROM chat_log ORDER BY asked_at ASC, rowid ASC"
            )
        except TursoError:
            logger.exception("Could not read the chat log from Turso")
            return []

        entries = []
        for row in rows:
            try:
                entries.append(json.loads(row["payload"]))
            except (KeyError, json.JSONDecodeError):
                continue
        return entries

    def page(
        self, limit: int = 50, offset: int = 0, search: str = ""
    ) -> tuple[list[dict[str, Any]], int]:
        """Newest first, optionally filtered by a free-text search."""
        entries = self.all_entries(newest_first=True)

        needle = search.strip().lower()
        if needle:
            fields = ("question", "answer", "name", "email", "company", "ip", "city", "country")
            entries = [
                entry
                for entry in entries
                if any(needle in str(entry.get(field, "")).lower() for field in fields)
            ]

        total = len(entries)
        return entries[offset : offset + limit], total

    def stats(self) -> dict[str, Any]:
        entries = self.all_entries()
        visitors = {entry.get("visitor_id") for entry in entries if entry.get("visitor_id")}
        identified = [entry for entry in entries if entry.get("email") or entry.get("name")]
        first = entries[0].get("asked_at") if entries else None
        last = entries[-1].get("asked_at") if entries else None
        return {
            "total_questions": len(entries),
            "unique_visitors": len(visitors),
            "identified_visitors": len({e.get("email") or e.get("name") for e in identified}),
            "flagged": sum(1 for e in entries if e.get("kind") == "abuse"),
            "first_asked_at": first,
            "last_asked_at": last,
            "xlsx_available": XLSX_AVAILABLE,
            "storage_path": "Turso (libSQL)" if self._turso else str(self._path),
        }

    # -- clearing (owner only) ---------------------------------------------

    def clear(self) -> int:
        """Delete every logged question. Returns how many were removed.

        Also stamps the reset marker, so returning visitors are asked for their
        name and email again instead of riding on a record that no longer exists.
        """
        with self._lock:
            removed = self._count
            if self._turso:
                try:
                    self._turso.execute("DELETE FROM chat_log")
                except TursoError:
                    logger.exception("Could not clear the Turso chat log")
                    raise
            try:
                if self._path.exists():
                    self._path.unlink()
            except OSError:
                logger.exception("Could not delete the chat log file")
                if not self._turso:
                    raise
            self._count = 0
            self._reset_at = time.time()
            self._write_reset_at(self._reset_at)
        logger.warning("Chat log cleared by the owner — %d entries deleted", removed)
        return removed

    # -- export ------------------------------------------------------------

    def to_rows(self) -> list[dict[str, Any]]:
        """Flatten the stored entries into spreadsheet-ready rows."""
        rows: list[dict[str, Any]] = []
        for index, entry in enumerate(self.all_entries(), start=1):
            asked_at = entry.get("asked_at", "")
            local = _to_ist(asked_at)
            row = {key: _cell(entry.get(key, "")) for _, key, _ in COLUMNS}
            row["row"] = index
            row["kind"] = KIND_LABELS.get(entry.get("kind", "chat"), entry.get("kind", ""))
            row["date_ist"] = local.strftime("%Y-%m-%d") if local else ""
            row["time_ist"] = local.strftime("%H:%M:%S") if local else ""
            row["asked_at"] = asked_at
            rows.append(row)
        return rows

    def to_xlsx(self) -> bytes:
        """Render the log as a real .xlsx workbook."""
        if not XLSX_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Chat questions"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="4F46E5")

        sheet.append([header for header, _, _ in COLUMNS])
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", horizontal="left")
        sheet.row_dimensions[1].height = 22

        for row in self.to_rows():
            sheet.append([row.get(key, "") for _, key, _ in COLUMNS])

        for index, (_, _, width) in enumerate(COLUMNS, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

        # Long free text stays readable instead of bleeding across the sheet.
        wrapped = {"question", "answer", "user_agent"}
        for index, (_, key, _) in enumerate(COLUMNS, start=1):
            if key not in wrapped:
                continue
            for cell in sheet[get_column_letter(index)][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        sheet.freeze_panes = "E2"  # keep #/date/time/type visible while scrolling
        sheet.auto_filter.ref = sheet.dimensions

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def to_csv(self) -> bytes:
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer)
        writer.writerow([header for header, _, _ in COLUMNS])
        for row in self.to_rows():
            writer.writerow([row.get(key, "") for _, key, _ in COLUMNS])
        # BOM so Excel opens UTF-8 (Hindi questions included) without mojibake.
        return buffer.getvalue().encode("utf-8-sig")


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------


def _cell(value: Any) -> Any:
    """Excel rejects some values outright; coerce them to something printable."""
    if value is None:
        return ""
    if isinstance(value, (int, float, str)):
        # 32767 is Excel's hard per-cell character limit.
        return value[:32000] if isinstance(value, str) and len(value) > 32000 else value
    if isinstance(value, (list, tuple)):
        return ", ".join(str(item) for item in value)
    return str(value)


def _to_ist(iso_timestamp: str) -> datetime | None:
    try:
        return datetime.fromisoformat(iso_timestamp.replace("Z", "+00:00")).astimezone(IST)
    except (AttributeError, ValueError):
        return None


_BROWSERS = [
    ("Edge", r"Edg[A-Z]?/"),
    ("Opera", r"OPR/|Opera"),
    ("Samsung Internet", r"SamsungBrowser"),
    ("Brave", r"Brave"),
    ("Chrome", r"Chrome/|CriOS"),
    ("Firefox", r"Firefox/|FxiOS"),
    ("Safari", r"Safari/"),
]

_PLATFORMS = [
    ("Android", r"Android"),
    ("iOS", r"iPhone|iPad|iPod"),
    ("Windows", r"Windows NT"),
    ("macOS", r"Mac OS X|Macintosh"),
    ("Linux", r"Linux|X11"),
]


def parse_user_agent(user_agent: str) -> dict[str, str]:
    """Best-effort browser/OS/device split — enough to know who is visiting."""
    if not user_agent:
        return {"browser": "", "os": "", "device": ""}

    browser = next((name for name, pattern in _BROWSERS if re.search(pattern, user_agent)), "Other")
    platform = next(
        (name for name, pattern in _PLATFORMS if re.search(pattern, user_agent)), "Unknown"
    )

    if re.search(r"iPad|Tablet", user_agent):
        device = "Tablet"
    elif re.search(r"Mobi|Android|iPhone", user_agent):
        device = "Mobile"
    elif re.search(r"bot|crawler|spider|curl|python-requests", user_agent, re.IGNORECASE):
        device = "Bot"
    else:
        device = "Desktop"

    return {"browser": browser, "os": platform, "device": device}
